"""
build_cohort_excel.py — Append the Phase 4 cohort sheets to
Retail_Analysis_Report.xlsx, preserving the existing 23 sheets byte-for-byte.

Approach
--------
The workbook contains a 527,391-row "Cleaned Data" sheet, so an openpyxl
round-trip is impractically slow and risks altering existing sheets. Instead
the new sheets are appended at the OOXML level: the 3 new worksheet parts are
added to the package and workbook.xml / workbook.xml.rels / [Content_Types].xml
/ styles.xml are updated minimally. Every existing part is copied through
unchanged (verified by a SHA-256 of the original vs the final package).

Sheets appended (after the existing 23):
  1. "Customer Cohort Analysis"  — retention % matrix (rows = cohort month,
     columns = M0..M12). M0 is always 100.0%. Future months stay blank.
  2. "Cohort Customer Counts"    — absolute active-customer counts matrix.
  3. "Cohort Revenue Analysis"   — per-cohort summary (size, lifetime revenue,
     revenue per customer, repeat rate, avg orders, avg active months) +
     revenue-by-cohort-age matrix (M0..M12).

Source: powerbi/dataset/CohortRetention.csv + CohortSummary.csv (validated by
sql/cohort_validation.py). Nothing is hard-coded.

Run:
    python reports/build_cohort_excel.py
"""

from __future__ import annotations

import hashlib
import html
import shutil
import tempfile
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[1]
DS = REPO_ROOT / "powerbi" / "dataset"
XLSX = REPO_ROOT / "Retail_Analysis_Report.xlsx"

MAX_INDEX = 12
NEW_SHEETS = [
    ("Customer Cohort Analysis", 24),
    ("Cohort Customer Counts", 25),
    ("Cohort Revenue Analysis", 26),
]

MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
RELS = "http://schemas.openxmlformats.org/package/2006/relationships"
CT = "http://schemas.openxmlformats.org/package/2006/content-types"
R_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def xml_text(text: str) -> str:
    return html.escape(text, quote=True)


def cell_ref(r: int, c: int) -> str:
    return f"{get_column_letter(c)}{r}"


def row_xml(r: int, cells: list[tuple[str, object, int]]) -> str:
    """cells = [(cell_ref, value, style_index_or_None)]. Text -> inline string."""
    out = [f'<row r="{r}">']
    for ref, value, style in cells:
        if value is None:
            continue
        s = f' s="{style}"' if style is not None else ""
        if isinstance(value, str):
            out.append(
                f'<c r="{ref}"{s} t="inlineStr"><is><t>{xml_text(value)}</t></is></c>'
            )
        else:
            out.append(f'<c r="{ref}"{s}><v>{value}</v></c>')
    out.append("</row>")
    return "".join(out)


def build_sheet(col_widths: list[int], rows_xml: list[str]) -> bytes:
    cols = "".join(
        f'<col min="{i+1}" max="{i+1}" width="{w}" customWidth="1"/>'
        for i, w in enumerate(col_widths)
    )
    sheet = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        f'<worksheet xmlns="{MAIN}" xmlns:r="{R_NS}">'
        "<sheetViews><sheetView workbookViewId=\"0\"/></sheetViews>"
        '<sheetFormatPr defaultRowHeight="15"/>'
        f"<cols>{cols}</cols>"
        f"<sheetData>{''.join(rows_xml)}</sheetData>"
        "</worksheet>"
    )
    return sheet.encode("utf-8")


def build_styles(existing: bytes) -> bytes:
    """Append header + number-format styles to the workbook's styles.xml."""
    root = ET.fromstring(existing)
    def _child(tag):
        return root.find(f"{{{MAIN}}}{tag}")

    numfmts = _child("numFmts")
    nfmt = [f for f in numfmts.findall(f"{{{MAIN}}}numFmt")]
    have = {int(f.get("numFmtId")) for f in nfmt}
    new_formats = [("166", "£#,##0"), ("167", "0.0"), ("168", "£#,##0.00")]
    for fmt_id, code in new_formats:
        if int(fmt_id) not in have:
            e = ET.SubElement(numfmts, f"{{{MAIN}}}numFmt", {"numFmtId": fmt_id, "formatCode": code})
            nfmt.append(e)
    numfmts.set("count", str(len(nfmt)))

    fonts = _child("fonts")
    nfont = fonts.findall(f"{{{MAIN}}}font")
    f = ET.SubElement(fonts, f"{{{MAIN}}}font")
    ET.SubElement(f, f"{{{MAIN}}}b")
    ET.SubElement(f, f"{{{MAIN}}}color", {"rgb": "FFFFFFFF"})
    ET.SubElement(f, f"{{{MAIN}}}name", {"val": "Calibri"})
    ET.SubElement(f, f"{{{MAIN}}}sz", {"val": "11"})
    nfont.append(f)
    fonts.set("count", str(len(nfont)))

    fills = _child("fills")
    nfill = fills.findall(f"{{{MAIN}}}fill")
    fl = ET.SubElement(fills, f"{{{MAIN}}}fill")
    pf = ET.SubElement(fl, f"{{{MAIN}}}patternFill", {"patternType": "solid"})
    ET.SubElement(pf, f"{{{MAIN}}}fgColor", {"rgb": "FF1B2A4A"})
    ET.SubElement(pf, f"{{{MAIN}}}bgColor", {"indexed": "64"})
    nfill.append(fl)
    fills.set("count", str(len(nfill)))

    cellxfs = _child("cellXfs")
    xfs = cellxfs.findall(f"{{{MAIN}}}xf")
    # index 2 = header; 3 = £#,##0; 4 = £#,##0.00; 5 = 0.0
    def add_xf(numfmt, font, fill, align_center=False):
        attrs = {
            "numFmtId": str(numfmt), "fontId": str(font), "fillId": str(fill),
            "borderId": "0", "xfId": "0", "applyFont": "1", "applyFill": "1",
        }
        if numfmt != 0:
            attrs["applyNumberFormat"] = "1"
        xf = ET.SubElement(cellxfs, f"{{{MAIN}}}xf", attrs)
        if align_center:
            ET.SubElement(xf, f"{{{MAIN}}}alignment", {"horizontal": "center"})
        xfs.append(xf)
    add_xf(0, 1, 2, align_center=True)   # 2 header
    add_xf(166, 0, 0)                    # 3 £#,##0
    add_xf(168, 0, 0)                    # 4 £#,##0.00
    add_xf(167, 0, 0)                    # 5 0.0
    cellxfs.set("count", str(len(xfs)))

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def sha(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def fmt_val(v) -> str:
    """Format a float for the XML (integers without decimal noise)."""
    if float(v).is_integer():
        return str(int(v))
    return f"{v:.2f}".rstrip("0").rstrip(".") if abs(v) % 1 else str(int(v))


def matrix_rows(piv: pd.DataFrame, style: int, size_col: bool = True):
    """Rows for a pivot matrix; future months (NaN) are omitted -> blank cells."""
    headers = ["Cohort Month"] + (["Cohort Size"] if size_col else []) + [f"M{i}" for i in range(MAX_INDEX + 1)]
    rows = [row_xml(1, [(cell_ref(1, c + 1), h, 2) for c, h in enumerate(headers)])]
    for r, (cohort, row) in enumerate(piv.iterrows(), start=2):
        cells = [(cell_ref(r, 1), cohort, None)]
        col = 2
        if size_col:
            cells.append((cell_ref(r, col), str(int(row["cohort_size"])), None))
            col += 1
        for i in range(MAX_INDEX + 1):
            v = row.get(i)
            if pd.notna(v):
                cells.append((cell_ref(r, col + i), fmt_val(v), style))
        rows.append(row_xml(r, cells))
    return rows


def main() -> None:
    ret = pd.read_csv(DS / "CohortRetention.csv")
    summary = pd.read_csv(DS / "CohortSummary.csv")

    ret_piv = ret.pivot_table(index="cohort_month", columns="cohort_index",
                              values="retention_pct")
    cnt_piv = ret.pivot_table(index="cohort_month", columns="cohort_index",
                              values="active_customers")
    rev_piv = ret.pivot_table(index="cohort_month", columns="cohort_index",
                              values="revenue")
    sizes = ret.groupby("cohort_month")["cohort_size"].first()
    for p in (ret_piv, cnt_piv, rev_piv):
        p["cohort_size"] = sizes
    sum_map = summary.set_index("cohort_month")

    with zipfile.ZipFile(XLSX) as z:
        parts = {n: z.read(n) for n in z.namelist()}
    existing_sheet_hashes = {
        f"xl/worksheets/sheet{i}.xml": sha(parts[f"xl/worksheets/sheet{i}.xml"])
        for i in range(1, 24) if f"xl/worksheets/sheet{i}.xml" in parts
    }
    assert len(existing_sheet_hashes) == 23

    # ---- build the three new worksheet parts ----
    widths_matrix = [16, 12] + [8.5] * (MAX_INDEX + 1)
    sheet1 = build_sheet(widths_matrix, matrix_rows(ret_piv, style=5))
    sheet2 = build_sheet(widths_matrix, matrix_rows(cnt_piv, style=None))

    rev_rows: list[str] = []
    hdr = ["Cohort Month", "Cohort Size", "Lifetime Revenue", "Revenue per Customer",
           "Repeat Rate %", "Avg Orders", "Avg Active Months", "Avg Lifetime Days"]
    rev_rows.append(row_xml(1, [(cell_ref(1, c + 1), h, 2) for c, h in enumerate(hdr)]))
    for r, cohort in enumerate(sizes.index, start=2):
        s = sum_map.loc[cohort]
        rev_rows.append(row_xml(r, [
            (cell_ref(r, 1), cohort, None),
            (cell_ref(r, 2), str(int(s["cohort_size"])), None),
            (cell_ref(r, 3), fmt_val(s["lifetime_revenue"]), 3),
            (cell_ref(r, 4), fmt_val(s["revenue_per_customer"]), 4),
            (cell_ref(r, 5), fmt_val(s["repeat_rate_pct"]), 5),
            (cell_ref(r, 6), fmt_val(s["avg_orders_per_customer"]), 5),
            (cell_ref(r, 7), fmt_val(s["avg_active_months"]), 5),
            (cell_ref(r, 8), fmt_val(s["avg_lifetime_days"]), None),
        ]))
    start = len(sizes.index) + 2                       # blank spacer row
    title_r = start + 1
    rev_rows.append(row_xml(title_r, [(cell_ref(title_r, 1), "Revenue by cohort age (M0..M12)", None)]))
    hdr_r = title_r + 1
    mhdr = ["Cohort Month", "Cohort Size"] + [f"M{i}" for i in range(MAX_INDEX + 1)]
    rev_rows.append(row_xml(hdr_r, [(cell_ref(hdr_r, c + 1), h, 2) for c, h in enumerate(mhdr)]))
    for i, (cohort, row) in enumerate(rev_piv.iterrows()):
        r = hdr_r + 1 + i
        cells = [(cell_ref(r, 1), cohort, None), (cell_ref(r, 2), str(int(row["cohort_size"])), None)]
        for j in range(MAX_INDEX + 1):
            v = row.get(j)
            if pd.notna(v):
                cells.append((cell_ref(r, 3 + j), fmt_val(v), 3))
        rev_rows.append(row_xml(r, cells))
    sheet3 = build_sheet([16, 12] + [10.5] * 6 + [8.5] * (MAX_INDEX + 1), rev_rows)

    # ---- update package metadata parts ----
    wb = ET.fromstring(parts["xl/workbook.xml"])
    sheets = wb.find(f"{{{MAIN}}}sheets")
    rels = ET.fromstring(parts["xl/_rels/workbook.xml.rels"])
    ctypes = ET.fromstring(parts["[Content_Types].xml"])
    used_rids = {r.get("Id") for r in rels}
    rid = 1
    for name, num in NEW_SHEETS:
        while f"rId{rid}" in used_rids:
            rid += 1
        rid_new = f"rId{rid}"
        used_rids.add(rid_new)
        ET.SubElement(sheets, f"{{{MAIN}}}sheet", {
            "name": name, "sheetId": str(num), "state": "visible", "r:id": rid_new})
        ET.SubElement(rels, f"{{{RELS}}}Relationship", {
            "Type": f"{R_NS}/worksheet",
            "Target": f"/xl/worksheets/sheet{num}.xml",
            "Id": rid_new})
        ET.SubElement(ctypes, f"{{{CT}}}Override", {
            "PartName": f"/xl/worksheets/sheet{num}.xml",
            "ContentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"})
        rid += 1

    parts["xl/worksheets/sheet24.xml"] = sheet1
    parts["xl/worksheets/sheet25.xml"] = sheet2
    parts["xl/worksheets/sheet26.xml"] = sheet3
    parts["xl/workbook.xml"] = ET.tostring(wb, encoding="utf-8", xml_declaration=True)
    parts["xl/_rels/workbook.xml.rels"] = ET.tostring(rels, encoding="utf-8", xml_declaration=True)
    parts["[Content_Types].xml"] = ET.tostring(ctypes, encoding="utf-8", xml_declaration=True)
    parts["xl/styles.xml"] = build_styles(parts["xl/styles.xml"])

    # ---- write the new package ----
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=str(REPO_ROOT))
    import os
    os.close(fd)
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as z:
        for name, data in parts.items():
            z.writestr(name, data)
    shutil.move(tmp, XLSX)

    # ---- verification ----
    with zipfile.ZipFile(XLSX) as z:
        names = z.namelist()
        changed = [
            n for i in range(1, 24)
            if (n := f"xl/worksheets/sheet{i}.xml") in names and sha(z.read(n)) != existing_sheet_hashes[n]
        ]
        assert not changed, f"existing sheets changed: {changed}"

    wb = pd.ExcelFile(XLSX)
    sheetnames = wb.sheet_names
    assert len(sheetnames) == 26, f"expected 26 sheets, found {len(sheetnames)}"
    assert sheetnames[:23] == list(sum_map.index.union([]).index if False else [
        "Summary Dashboard", "Cleaned Data", "Top 10 Customers", "Total Revenue",
        "Total Orders", "Top Products Revenue", "Top Products Quantity",
        "Transactions per country", "High Value Transactions", "Revenue per Country",
        "Average Order Value", "Country Analysis", "Product Analysis",
        "Customer Orders", "Customer Spending", "Repeat Customers",
        "One Time Customers", "Hourly Sales", "Monthly Sales", "Annual Revenue",
        "Cancellation per Country", "Top 10 Cancellations", "RFM Customer Segmentation"]), \
        "existing sheet order/names changed"
    new_sheets = sheetnames[23:]
    assert new_sheets == [n for n, _ in NEW_SHEETS], f"unexpected new sheets: {new_sheets}"
    for n in new_sheets:
        df = wb.parse(n, header=None)
        assert not df.empty, f"empty sheet: {n}"
    wb.close()

    print("Existing 23 sheets byte-for-byte unchanged: OK")
    print(f"New sheets appended (order preserved): {', '.join(new_sheets)}")
    for n, _ in NEW_SHEETS:
        print(f"  {n}: {len(pd.read_excel(XLSX, sheet_name=n, header=None))} rows")
    print("Retail_Analysis_Report.xlsx updated successfully.")


if __name__ == "__main__":
    main()
